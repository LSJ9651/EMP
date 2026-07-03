"""报表中心路由 — 日报/周报/月报 + 设备明细/告警历史导出

性能优化要点：
1. 电价策略一次性加载，消除 N+1 查询
2. 时段分类改为纯内存计算，不再每条记录查库
3. 移除 _get_daily_summary 中无用的 records 列表
4. 日报复用 _get_daily_summary 逻辑
"""
from datetime import datetime, timedelta
from io import BytesIO
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from models import Telemetry, Device, AlertRecord, TariffPolicy

router = APIRouter(prefix="/api/report-center", tags=["报表中心"])

# ──────────────────── 电价时段缓存 ────────────────────


def _build_tariff_classifier(db: Session):
    """构建时段分类器（一次查询，纯内存判断）"""
    tariffs = db.query(TariffPolicy).filter(TariffPolicy.is_active == True).all()

    rules = []
    for t in tariffs:
        start = t.start_time  # "HH:MM"
        end = t.end_time
        rules.append((start, end, t.period_name, t.price_per_kwh))

    # 价格映射
    price_map = {t.period_name: t.price_per_kwh for t in tariffs}

    def classify(ts: datetime):
        """根据时间戳分类时段（纯内存，无DB查询）"""
        t_str = ts.strftime("%H:%M")
        for start, end, period, _ in rules:
            if start <= end:
                if start <= t_str <= end:
                    return period
            else:  # 跨午夜时段
                if t_str >= start or t_str <= end:
                    return period
        return "平段"

    return classify, price_map


# ──────────────────── 核心聚合函数 ────────────────────


def _get_daily_summary(start: datetime, end: datetime, db: Session):
    """获取指定时间范围内的每日汇总数据（优化版：消除N+1查询）"""
    # 一次性加载电价分类器和价格映射
    classify, price_map = _build_tariff_classifier(db)

    records = (
        db.query(Telemetry)
        .filter(Telemetry.timestamp >= start, Telemetry.timestamp <= end)
        .all()
    )

    daily = {}
    for r in records:
        day = r.timestamp.strftime("%Y-%m-%d")
        if day not in daily:
            daily[day] = {
                "total_energy_kwh": 0.0,
                "peak_energy": 0.0,
                "flat_energy": 0.0,
                "valley_energy": 0.0,
                "peak_cost": 0.0,
                "flat_cost": 0.0,
                "valley_cost": 0.0,
                "max_power": 0.0,
            }
        d = daily[day]
        d["total_energy_kwh"] += r.energy_kwh
        if r.power > d["max_power"]:
            d["max_power"] = r.power

        # 纯内存时段分类，无DB查询
        period = classify(r.timestamp)
        price = price_map.get(period, 0.8)
        cost = r.energy_kwh * price

        if period == "高峰":
            d["peak_energy"] += r.energy_kwh
            d["peak_cost"] += cost
        elif period == "低谷":
            d["valley_energy"] += r.energy_kwh
            d["valley_cost"] += cost
        else:
            d["flat_energy"] += r.energy_kwh
            d["flat_cost"] += cost

    result = []
    for day in sorted(daily.keys()):
        d = daily[day]
        total_cost = d["peak_cost"] + d["flat_cost"] + d["valley_cost"]
        result.append({
            "date": day,
            "total_energy_kwh": round(d["total_energy_kwh"], 2),
            "peak_energy_kwh": round(d["peak_energy"], 2),
            "flat_energy_kwh": round(d["flat_energy"], 2),
            "valley_energy_kwh": round(d["valley_energy"], 2),
            "max_power_kw": round(d["max_power"], 2),
            "co2_kg": round(d["total_energy_kwh"] * 0.5703, 2),
            "cost_yuan": round(total_cost, 2),
        })
    return result


# ──────────────────── API 路由 ────────────────────


@router.get("/daily")
def daily_report(
    date: str = Query(None, description="查询日期，格式 YYYY-MM-DD，默认今天"),
    db: Session = Depends(get_db),
):
    """日报：指定日期总能耗、峰平谷用电量、碳排放、最大功率、成本"""
    if date:
        try:
            today = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)

    # 复用 _get_daily_summary（消除重复代码），取当天唯一一条结果
    result = _get_daily_summary(today, tomorrow - timedelta(seconds=1), db)
    data = result[0] if result else {
        "date": today.strftime("%Y-%m-%d"),
        "total_energy_kwh": 0,
        "peak_energy_kwh": 0,
        "flat_energy_kwh": 0,
        "valley_energy_kwh": 0,
        "max_power_kw": 0,
        "co2_kg": 0,
        "cost_yuan": 0,
    }
    return {"code": 200, "data": data, "message": "success"}


@router.get("/weekly")
def weekly_report(db: Session = Depends(get_db)):
    """周报：过去7天每日汇总"""
    end = datetime.now().replace(hour=23, minute=59, second=59)
    start = (end - timedelta(days=7)).replace(hour=0, minute=0, second=0)
    result = _get_daily_summary(start, end, db)
    return {"code": 200, "data": result, "message": "success"}


@router.get("/monthly")
def monthly_report(db: Session = Depends(get_db)):
    """月报：过去30天每日汇总"""
    end = datetime.now().replace(hour=23, minute=59, second=59)
    start = (end - timedelta(days=30)).replace(hour=0, minute=0, second=0)
    result = _get_daily_summary(start, end, db)
    return {"code": 200, "data": result, "message": "success"}


# ──────────────────── 报表导出接口 ────────────────────


def _export_report_to_excel(data, sheet_name, filename_prefix):
    """通用报表导出函数"""
    try:
        import openpyxl
    except ImportError:
        return {"code": 500, "data": None, "message": "openpyxl 未安装"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if isinstance(data, dict):
        # 日报单条数据
        headers = ["指标", "数值"]
        ws.append(headers)
        ws.append(["日期", data.get("date", "")])
        ws.append(["总能耗(kWh)", data.get("total_energy_kwh", 0)])
        ws.append(["峰时用电(kWh)", data.get("peak_energy_kwh", 0)])
        ws.append(["平时用电(kWh)", data.get("flat_energy_kwh", 0)])
        ws.append(["谷时用电(kWh)", data.get("valley_energy_kwh", 0)])
        ws.append(["最大功率(kW)", data.get("max_power_kw", 0)])
        ws.append(["碳排放(kgCO₂)", data.get("co2_kg", 0)])
        ws.append(["预估成本(元)", data.get("cost_yuan", 0)])
    else:
        # 周报/月报多条数据
        headers = ["日期", "总能耗(kWh)", "峰时用电(kWh)", "平时用电(kWh)", "谷时用电(kWh)", 
                   "最大功率(kW)", "碳排放(kgCO₂)", "预估成本(元)"]
        ws.append(headers)
        for row in data:
            ws.append([
                row.get("date", ""),
                row.get("total_energy_kwh", 0),
                row.get("peak_energy_kwh", 0),
                row.get("flat_energy_kwh", 0),
                row.get("valley_energy_kwh", 0),
                row.get("max_power_kw", 0),
                row.get("co2_kg", 0),
                row.get("cost_yuan", 0),
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/daily/export")
def export_daily_report(
    date: str = Query(None, description="查询日期，格式 YYYY-MM-DD，默认今天"),
    db: Session = Depends(get_db),
):
    """日报导出 Excel"""
    if date:
        try:
            today = datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    else:
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    tomorrow = today + timedelta(days=1)

    result = _get_daily_summary(today, tomorrow - timedelta(seconds=1), db)
    data = result[0] if result else {
        "date": today.strftime("%Y-%m-%d"),
        "total_energy_kwh": 0,
        "peak_energy_kwh": 0,
        "flat_energy_kwh": 0,
        "valley_energy_kwh": 0,
        "max_power_kw": 0,
        "co2_kg": 0,
        "cost_yuan": 0,
    }
    return _export_report_to_excel(data, "能耗日报", "daily_report")


@router.get("/weekly/export")
def export_weekly_report(db: Session = Depends(get_db)):
    """周报导出 Excel"""
    end = datetime.now().replace(hour=23, minute=59, second=59)
    start = (end - timedelta(days=7)).replace(hour=0, minute=0, second=0)
    result = _get_daily_summary(start, end, db)
    return _export_report_to_excel(result, "能耗周报", "weekly_report")


@router.get("/monthly/export")
def export_monthly_report(db: Session = Depends(get_db)):
    """月报导出 Excel"""
    end = datetime.now().replace(hour=23, minute=59, second=59)
    start = (end - timedelta(days=30)).replace(hour=0, minute=0, second=0)
    result = _get_daily_summary(start, end, db)
    return _export_report_to_excel(result, "能耗月报", "monthly_report")


@router.get("/devices/export")
def export_devices(
    start: str = Query(None, description="开始时间"),
    end: str = Query(None, description="结束时间"),
    db: Session = Depends(get_db),
):
    """设备明细导出 Excel"""
    try:
        import openpyxl
    except ImportError:
        return {"code": 500, "data": None, "message": "openpyxl 未安装"}

    try:
        st = datetime.fromisoformat(start) if start else datetime.now() - timedelta(days=7)
        et = datetime.fromisoformat(end) if end else datetime.now()
    except ValueError:
        return {"code": 400, "data": None, "message": "时间格式错误"}

    records = (
        db.query(Telemetry)
        .filter(Telemetry.timestamp >= st, Telemetry.timestamp <= et)
        .order_by(Telemetry.timestamp.asc())
        .all()
    )

    devices = {d.id: d.name for d in db.query(Device).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备能耗明细"

    headers = ["序号", "设备名称", "时间", "电压(V)", "电流(A)", "功率(kW)",
               "电度(kWh)", "功率因数", "温度(°C)", "状态码"]
    ws.append(headers)

    for i, r in enumerate(records, 1):
        ws.append([
            i,
            devices.get(r.device_id, f"设备{r.device_id}"),
            r.timestamp.strftime("%Y-%m-%d %H:%M:%S") if r.timestamp else "",
            r.voltage,
            r.current,
            r.power,
            r.energy_kwh,
            r.power_factor,
            r.temperature,
            r.status_code,
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"device_detail_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/alerts/export")
def export_alerts(
    start: str = Query(None, description="开始时间"),
    end: str = Query(None, description="结束时间"),
    db: Session = Depends(get_db),
):
    """告警历史导出 Excel"""
    try:
        import openpyxl
    except ImportError:
        return {"code": 500, "data": None, "message": "openpyxl 未安装"}

    try:
        st = datetime.fromisoformat(start) if start else datetime.now() - timedelta(days=30)
        et = datetime.fromisoformat(end) if end else datetime.now()
    except ValueError:
        return {"code": 400, "data": None, "message": "时间格式错误"}

    records = (
        db.query(AlertRecord, Device.name)
        .join(Device, AlertRecord.device_id == Device.id)
        .filter(AlertRecord.alert_time >= st, AlertRecord.alert_time <= et)
        .order_by(AlertRecord.alert_time.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "告警历史"

    headers = ["序号", "设备名称", "告警时间", "参数类型", "触发值", "阈值",
               "告警信息", "严重程度", "是否处理", "处理人", "处理措施", "处理时间"]
    ws.append(headers)

    for i, (r, name) in enumerate(records, 1):
        ws.append([
            i,
            name,
            r.alert_time.strftime("%Y-%m-%d %H:%M:%S") if r.alert_time else "",
            r.param_type,
            r.value,
            r.threshold_value,
            r.message,
            r.severity,
            "是" if r.is_resolved else "否",
            r.handler or "",
            r.measure or "",
            r.resolved_at.strftime("%Y-%m-%d %H:%M:%S") if r.resolved_at else "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"alert_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
